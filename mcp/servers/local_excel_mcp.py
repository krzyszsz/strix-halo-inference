#!/usr/bin/env python
import math
import os
from pathlib import Path
from typing import Any

from fastmcp import FastMCP
from openpyxl import Workbook, load_workbook

mcp = FastMCP("local-excel")


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _load_or_create(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    return wb


def _get_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


@mcp.tool
def excel_new(path: str) -> str:
    """Create a new workbook at path and return the saved path."""
    p = Path(path)
    _ensure_parent(p)
    wb = Workbook()
    wb.save(p)
    return str(p)


@mcp.tool
def excel_set_cell(path: str, sheet: str, cell: str, value: Any) -> str:
    """Set a cell value and save the workbook."""
    p = Path(path)
    _ensure_parent(p)
    wb = _load_or_create(p)
    ws = _get_sheet(wb, sheet)
    ws[cell] = value
    wb.save(p)
    return f"{sheet}!{cell}={value}"


@mcp.tool
def excel_get_cell(path: str, sheet: str, cell: str) -> Any:
    """Get a cell value from a workbook."""
    p = Path(path)
    wb = load_workbook(p, data_only=True)
    ws = _get_sheet(wb, sheet)
    return ws[cell].value


@mcp.tool
def excel_sum_range(path: str, sheet: str, cell_range: str) -> float:
    """Compute a numeric sum over a cell range and return it."""
    p = Path(path)
    wb = load_workbook(p, data_only=True)
    ws = _get_sheet(wb, sheet)
    total = 0.0
    for row in ws[cell_range]:
        for cell in row:
            if isinstance(cell.value, (int, float)) and not math.isnan(cell.value):
                total += float(cell.value)
    return total


if __name__ == "__main__":
    transport = os.getenv("MCP_TRANSPORT", "stdio")
    host = os.getenv("MCP_HOST", "127.0.0.1")
    port = int(os.getenv("MCP_PORT", "8017"))
    if transport in {"sse", "streamable-http"}:
        mcp.run(transport=transport, host=host, port=port, show_banner=False)
    else:
        mcp.run(show_banner=False)
